import json
import math
from typing import Dict, Any, List
import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.app.models import ProductSPU, ProductSKU, ChannelPricing
from backend.app.services.products import validate_channel_pricing_values

def _clean_val(v):
    if pd.isna(v):
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    return v

def _parse_price(v):
    v = _clean_val(v)
    if v is None:
        return None
    try:
        return int(float(v) * 100)
    except ValueError:
        return None

def _parse_date(v):
    v = _clean_val(v)
    if v is None:
        return None
    try:
        return pd.to_datetime(v).to_pydatetime()
    except Exception:
        return None


def _first_val(row, *names):
    for name in names:
        value = _clean_val(row.get(name))
        if value is not None:
            return value
    return None

def preview_excel_import(file_path: str, session: Session) -> Dict[str, Any]:
    """Parse the Excel file and return a preview of changes (New, Conflict)."""
    xls = pd.ExcelFile(file_path)
    
    result = {
        "spu": {"new": [], "conflict": []},
        "sku": {"new": [], "conflict": []},
        "pricing": {"new": [], "conflict": []}
    }
    
    # 1. Parse SPU
    spu_sheet = next((name for name in ["SPU_物料基础信息", "SPU_产品基础信息"] if name in xls.sheet_names), None)
    if spu_sheet:
        df_spu = pd.read_excel(xls, spu_sheet)
        for _, row in df_spu.iterrows():
            spu_id = _clean_val(row.get("SPU_ID"))
            if not spu_id:
                continue
            
            item = {
                "spu_id": str(spu_id),
                "name": str(_first_val(row, "物料名称(中文)", "产品名称(中文)", "商品名称(中文)") or ""),
                "name_en": _first_val(row, "物料名称(英文)", "产品名称(英文)", "商品名称(英文)"),
                "product_line": _first_val(row, "物料线", "产品线", "商品线"),
                "product_type": _first_val(row, "物料类型", "产品类型", "商品类型"),
                "brand": _clean_val(row.get("品牌")),
                "positioning": _first_val(row, "物料定位", "产品定位", "商品定位"),
                "launch_time": _clean_val(row.get("上市时间")),
                "lifecycle": _clean_val(row.get("生命周期")),
                "core_selling_points": _clean_val(row.get("核心卖点")),
                "use_cases": _clean_val(row.get("应用场景")),
                "target_users": _clean_val(row.get("目标用户")),
            }
            
            existing = session.execute(select(ProductSPU).where(ProductSPU.spu_id == str(spu_id))).scalars().first()
            if existing:
                result["spu"]["conflict"].append(item)
            else:
                result["spu"]["new"].append(item)

    # 2. Parse SKU
    if "SKU_销售单元" in xls.sheet_names:
        df_sku = pd.read_excel(xls, "SKU_销售单元")
        for _, row in df_sku.iterrows():
            sku_id = _clean_val(row.get("SKU_ID"))
            spu_id = _clean_val(row.get("SPU_ID"))
            if not sku_id or not spu_id:
                continue
            
            item = {
                "sku_id": str(sku_id),
                "spu_id": str(spu_id),
                "model": _clean_val(row.get("型号")),
                "version": _clean_val(row.get("版本")),
                "package_contents": _clean_val(row.get("套装内容")),
                "barcode": str(_clean_val(row.get("条码"))) if _clean_val(row.get("条码")) else None,
                "weight": _clean_val(row.get("重量")),
                "dimensions": _clean_val(row.get("尺寸")),
                "color": _clean_val(row.get("颜色")),
                "cost_price": _parse_price(row.get("成本价")),
                "msrp": _parse_price(row.get("建议零售价")),
            }
            
            existing = session.execute(select(ProductSKU).where(ProductSKU.sku_id == str(sku_id))).scalars().first()
            if existing:
                result["sku"]["conflict"].append(item)
            else:
                result["sku"]["new"].append(item)

    # 3. Parse Pricing
    if "渠道映射_增强" in xls.sheet_names:
        df_pricing = pd.read_excel(xls, "渠道映射_增强")
        for _, row in df_pricing.iterrows():
            sku_id = _clean_val(row.get("SKU_ID"))
            channel = _clean_val(row.get("渠道名称"))
            if not sku_id or not channel:
                continue
            
            item = {
                "sku_id": str(sku_id),
                "channel": str(channel),
                "channel_sku_id": _clean_val(row.get("渠道SKU_ID")),
                "listing_id": _clean_val(row.get("Listing_ID")),
                "status": _clean_val(row.get("上架状态")),
                "tier_a_price": _parse_price(row.get("日常售价")),
                "tier_b_price": _parse_price(row.get("一级促销价（如大促）")),
                "tier_c_price": _parse_price(row.get("二级促销价（如日常折扣）")),
                "map_price": _parse_price(row.get("最低限价")),
                "max_price": _parse_price(row.get("最高限价")),
                "promo_start_time": _parse_date(row.get("促销开始时间")),
                "promo_end_time": _parse_date(row.get("促销结束时间")),
                "currency": _clean_val(row.get("货币")) or "CNY",
                "stock_quantity": _clean_val(row.get("渠道库存")),
                "manager": _clean_val(row.get("渠道负责人")),
            }
            
            sku_db = session.execute(select(ProductSKU).where(ProductSKU.sku_id == str(sku_id))).scalars().first()
            sku_uuid = sku_db.id if sku_db else None
            
            if sku_uuid:
                existing = session.execute(select(ChannelPricing).where(ChannelPricing.sku_uuid == sku_uuid, ChannelPricing.channel == str(channel))).scalars().first()
                if existing:
                    result["pricing"]["conflict"].append(item)
                else:
                    result["pricing"]["new"].append(item)
            else:
                # Treat as new, but it will require the SKU to be imported first
                result["pricing"]["new"].append(item)

    return result


def confirm_excel_import(data: Dict[str, Any], session: Session) -> Dict[str, int]:
    """Execute the import logic using the accepted preview data."""
    counts = {"spu": 0, "sku": 0, "pricing": 0}
    
    # 1. SPU
    for spu_data in data.get("spu", []):
        spu_id = spu_data["spu_id"]
        existing = session.execute(select(ProductSPU).where(ProductSPU.spu_id == spu_id)).scalars().first()
        
        extended_info = {
            "core_selling_points": spu_data.get("core_selling_points"),
            "use_cases": spu_data.get("use_cases"),
            "target_users": spu_data.get("target_users"),
        }
        
        if existing:
            existing.name = spu_data.get("name", existing.name)
            existing.name_en = spu_data.get("name_en", existing.name_en)
            existing.product_line = spu_data.get("product_line", existing.product_line)
            existing.product_type = spu_data.get("product_type", existing.product_type)
            existing.brand = spu_data.get("brand", existing.brand)
            existing.positioning = spu_data.get("positioning", existing.positioning)
            existing.launch_time = str(spu_data.get("launch_time")) if spu_data.get("launch_time") else existing.launch_time
            existing.lifecycle = spu_data.get("lifecycle", existing.lifecycle)
            existing.extended_info_json = json.dumps(extended_info)
        else:
            new_spu = ProductSPU(
                spu_id=spu_id,
                name=spu_data.get("name", spu_id),
                name_en=spu_data.get("name_en"),
                product_line=spu_data.get("product_line"),
                product_type=spu_data.get("product_type"),
                brand=spu_data.get("brand"),
                positioning=spu_data.get("positioning"),
                launch_time=str(spu_data.get("launch_time")) if spu_data.get("launch_time") else None,
                lifecycle=spu_data.get("lifecycle"),
                extended_info_json=json.dumps(extended_info)
            )
            session.add(new_spu)
        counts["spu"] += 1
    
    session.flush()

    # 2. SKU
    for sku_data in data.get("sku", []):
        sku_id = sku_data["sku_id"]
        spu_id = sku_data["spu_id"]
        
        # find spu
        spu = session.execute(select(ProductSPU).where(ProductSPU.spu_id == spu_id)).scalars().first()
        if not spu:
            continue
            
        existing = session.execute(select(ProductSKU).where(ProductSKU.sku_id == sku_id)).scalars().first()
        
        attributes = {
            "weight": sku_data.get("weight"),
            "dimensions": sku_data.get("dimensions"),
            "color": sku_data.get("color"),
            "package_contents": sku_data.get("package_contents")
        }
        
        if existing:
            existing.model = sku_data.get("model", existing.model)
            existing.version = sku_data.get("version", existing.version)
            existing.barcode = str(sku_data.get("barcode")) if sku_data.get("barcode") else existing.barcode
            existing.cost_price = sku_data.get("cost_price", existing.cost_price)
            existing.msrp = sku_data.get("msrp", existing.msrp)
            existing.attributes_json = json.dumps(attributes)
        else:
            new_sku = ProductSKU(
                spu_uuid=spu.id,
                sku_id=sku_id,
                model=sku_data.get("model"),
                version=sku_data.get("version"),
                barcode=str(sku_data.get("barcode")) if sku_data.get("barcode") else None,
                cost_price=sku_data.get("cost_price"),
                msrp=sku_data.get("msrp"),
                attributes_json=json.dumps(attributes)
            )
            session.add(new_sku)
        counts["sku"] += 1
    
    session.flush()

    # 3. Pricing
    for pricing_data in data.get("pricing", []):
        sku_id = pricing_data["sku_id"]
        channel = pricing_data["channel"]
        
        sku = session.execute(select(ProductSKU).where(ProductSKU.sku_id == sku_id)).scalars().first()
        if not sku:
            continue
            
        existing = session.execute(select(ChannelPricing).where(
            ChannelPricing.sku_uuid == sku.id,
            ChannelPricing.channel == channel
        )).scalars().first()
        
        if existing:
            validate_channel_pricing_values(
                tier_a_price=pricing_data.get("tier_a_price", existing.tier_a_price),
                tier_b_price=pricing_data.get("tier_b_price", existing.tier_b_price),
                tier_c_price=pricing_data.get("tier_c_price", existing.tier_c_price),
                map_price=pricing_data.get("map_price", existing.map_price),
                max_price=pricing_data.get("max_price", existing.max_price),
                promo_start_time=pricing_data.get("promo_start_time", existing.promo_start_time),
                promo_end_time=pricing_data.get("promo_end_time", existing.promo_end_time),
            )
            existing.channel_sku_id = pricing_data.get("channel_sku_id", existing.channel_sku_id)
            existing.listing_id = pricing_data.get("listing_id", existing.listing_id)
            existing.status = pricing_data.get("status", existing.status)
            existing.tier_a_price = pricing_data.get("tier_a_price", existing.tier_a_price)
            existing.tier_b_price = pricing_data.get("tier_b_price", existing.tier_b_price)
            existing.tier_c_price = pricing_data.get("tier_c_price", existing.tier_c_price)
            existing.map_price = pricing_data.get("map_price", existing.map_price)
            existing.max_price = pricing_data.get("max_price", existing.max_price)
            if "promo_start_time" in pricing_data:
                existing.promo_start_time = pricing_data["promo_start_time"]
            if "promo_end_time" in pricing_data:
                existing.promo_end_time = pricing_data["promo_end_time"]
            existing.currency = pricing_data.get("currency", existing.currency)
            existing.stock_quantity = pricing_data.get("stock_quantity", existing.stock_quantity)
            existing.manager = pricing_data.get("manager", existing.manager)
        else:
            validate_channel_pricing_values(
                tier_a_price=pricing_data.get("tier_a_price"),
                tier_b_price=pricing_data.get("tier_b_price"),
                tier_c_price=pricing_data.get("tier_c_price"),
                map_price=pricing_data.get("map_price"),
                max_price=pricing_data.get("max_price"),
                promo_start_time=pricing_data.get("promo_start_time"),
                promo_end_time=pricing_data.get("promo_end_time"),
            )
            new_pricing = ChannelPricing(
                sku_uuid=sku.id,
                channel=channel,
                channel_sku_id=pricing_data.get("channel_sku_id"),
                listing_id=pricing_data.get("listing_id"),
                status=pricing_data.get("status"),
                tier_a_price=pricing_data.get("tier_a_price"),
                tier_b_price=pricing_data.get("tier_b_price"),
                tier_c_price=pricing_data.get("tier_c_price"),
                map_price=pricing_data.get("map_price"),
                max_price=pricing_data.get("max_price"),
                promo_start_time=pricing_data.get("promo_start_time"),
                promo_end_time=pricing_data.get("promo_end_time"),
                currency=pricing_data.get("currency") or "CNY",
                stock_quantity=pricing_data.get("stock_quantity"),
                manager=pricing_data.get("manager")
            )
            session.add(new_pricing)
        counts["pricing"] += 1
        
    return counts


def import_inventory_excel(file_path: str, session: Session) -> dict[str, Any]:
    import openpyxl
    from backend.app.models import ProductInventorySnapshot, now_utc
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Find the correct sheet
    sheet_name = next((name for name in ["海外库存总表", "Sheet1"] if name in wb.sheetnames), wb.sheetnames[0])
    sheet = wb[sheet_name]
    
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 文件是空的")
        
    header = [str(cell or "").strip() for cell in rows[0]]
    
    # Map headers to indices
    material_code_idx = -1
    english_name_idx = -1
    wh_idx = -1
    name_idx = -1
    qty_idx = -1
    in_transit_idx = -1
    warning_idx = -1
    model_idx = -1
    
    # Match headers
    for idx, h_name in enumerate(header):
        h_clean = h_name.replace("\n", " ").replace("/", " ").strip()
        h_clean_lower = h_clean.lower()
        
        if h_clean_lower == "sku":
            english_name_idx = idx
        elif "material code" in h_clean_lower or "料号" in h_clean_lower:
            material_code_idx = idx
        elif "仓库" in h_clean:
            wh_idx = idx
        elif "chinese description" in h_clean_lower or "中文品名" in h_clean or "品名" in h_clean:
            name_idx = idx
        elif "库存数量" in h_clean or "库存" in h_clean:
            if qty_idx == -1:
                qty_idx = idx
        elif "在途数量" in h_clean or "在途" in h_clean:
            in_transit_idx = idx
        elif "预警" in h_clean or "warning" in h_clean_lower:
            warning_idx = idx
        elif "型号" in h_clean or "model" in h_clean_lower:
            model_idx = idx

    # If any essential column is missing, try fallback positions
    if material_code_idx == -1:
        material_code_idx = 3 if len(header) > 3 else 0
    if english_name_idx == -1:
        english_name_idx = 1 if len(header) > 1 else 0
    if wh_idx == -1:
        wh_idx = 2
    if name_idx == -1:
        name_idx = 4
    if qty_idx == -1:
        qty_idx = 8
        
    # Read the data rows
    data_rows = rows[1:]
    synced_at = now_utc()
    
    # Delete existing inventory snapshots for warehouses present in the import data
    # (instead of deleting ALL warehouses, to avoid clearing other warehouses' data)
    warehouses_in_file = set()
    for row in data_rows:
        if len(row) > wh_idx and row[wh_idx] is not None:
            wh_val = str(row[wh_idx]).strip()
            if wh_val:
                warehouses_in_file.add(wh_val)
    if warehouses_in_file:
        for wh_code in warehouses_in_file:
            session.query(ProductInventorySnapshot).filter(
                ProductInventorySnapshot.warehouse_code == wh_code
            ).delete()

    # Aggregate by (material_code, warehouse_code) to prevent UniqueConstraint violations
    aggregated = {}
    for row in data_rows:
        if len(row) <= max(material_code_idx, wh_idx):
            continue
            
        material_code_val = row[material_code_idx]
        wh_val = row[wh_idx]
        if material_code_val is None or wh_val is None:
            continue
            
        material_code_str = str(material_code_val).strip()
        wh_str = str(wh_val).strip()
        if not material_code_str or not wh_str:
            continue
            
        # Parse quantities
        qty_val = 0.0
        if qty_idx < len(row) and row[qty_idx] is not None:
            try:
                qty_val = float(row[qty_idx])
            except ValueError:
                pass
                
        in_transit_val = 0.0
        if in_transit_idx != -1 and in_transit_idx < len(row) and row[in_transit_idx] is not None:
            try:
                in_transit_val = float(row[in_transit_idx])
            except ValueError:
                pass
                
        warning_str = "正常"
        if warning_idx != -1 and warning_idx < len(row) and row[warning_idx] is not None:
            warning_str = str(row[warning_idx]).strip()
            
        name_str = material_code_str
        if name_idx < len(row) and row[name_idx] is not None:
            name_str = str(row[name_idx]).strip()
            
        english_name_str = ""
        if english_name_idx != -1 and english_name_idx < len(row) and row[english_name_idx] is not None:
            english_name_str = str(row[english_name_idx]).strip()
            
        model_str = ""
        if model_idx != -1 and model_idx < len(row) and row[model_idx] is not None:
            model_str = str(row[model_idx]).strip()
            
        key = (material_code_str, wh_str)
        if key not in aggregated:
            aggregated[key] = {
                "material_name": name_str,
                "english_name": english_name_str,
                "qty": qty_val,
                "in_transit_qty": in_transit_val,
                "warning_status": warning_str,
                "model": model_str
            }
        else:
            # Aggregate quantities
            aggregated[key]["qty"] += qty_val
            aggregated[key]["in_transit_qty"] += in_transit_val
            if warning_str != "正常":
                aggregated[key]["warning_status"] = warning_str
            if model_str:
                aggregated[key]["model"] = model_str
            if english_name_str:
                aggregated[key]["english_name"] = english_name_str
                
    imported_count = 0
    import json
    for (material_code_str, wh_str), data in aggregated.items():
        payload = {
            "in_transit_qty": data["in_transit_qty"],
            "warning_status": data["warning_status"],
            "model": data["model"],
            "english_name": data["english_name"]
        }
        
        snapshot = ProductInventorySnapshot(
            material_code=material_code_str,
            material_name=data["material_name"],
            warehouse_code=wh_str,
            warehouse_name=wh_str,
            qty=data["qty"],
            base_qty=data["qty"],
            source_payload_json=json.dumps(payload),
            status="Active",
            synced_at=synced_at,
            created_at=synced_at,
            updated_at=synced_at
        )
        session.add(snapshot)
        imported_count += 1
        
    session.flush()
    return {
        "ok": True,
        "imported_count": imported_count
    }
