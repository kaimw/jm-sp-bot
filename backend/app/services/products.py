import json
import logging
import re
from difflib import SequenceMatcher
from datetime import datetime, timezone
from sqlalchemy.orm import Session
from sqlalchemy import exists, select

from backend.app.models import ProductInventorySnapshot, ProductSPU, ProductSKU, ChannelPricing, PromotionRule, SystemConfig, new_id
from backend.app.services.jsonutil import loads


logger = logging.getLogger(__name__)
FINISHED_CATEGORY = "成品"

def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def get_spu(session: Session, spu_id: str) -> ProductSPU | None:
    return session.query(ProductSPU).filter(ProductSPU.id == spu_id).first()


def finished_inventory_spu_exists():
    return exists().where(
        ProductInventorySnapshot.material_code == ProductSPU.spu_id,
    )


def filter_finished_inventory_spus(query):
    return query.filter(ProductSPU.category == FINISHED_CATEGORY).filter(finished_inventory_spu_exists())


def filter_finished_inventory_skus(query):
    return filter_finished_inventory_spus(query.join(ProductSPU, ProductSPU.id == ProductSKU.spu_uuid))


def ensure_finished_inventory_sku(session: Session, sku_uuid: str) -> ProductSKU:
    sku = (
        filter_finished_inventory_skus(session.query(ProductSKU))
        .filter(ProductSKU.id == sku_uuid)
        .one_or_none()
    )
    if sku is None:
        raise ValueError("SKU 必须属于成品库存中的物料，材料库存物料不能绑定 SKU/价格")
    return sku


def normalize_product_review_aliases(values) -> list[str]:
    if values is None:
        return []
    if isinstance(values, str):
        raw_values = re.split(r"[\n,，、;；]+", values)
    elif isinstance(values, list):
        raw_values = []
        for item in values:
            if isinstance(item, str):
                raw_values.extend(re.split(r"[\n,，、;；]+", str(item)))
    else:
        raw_values = []
    aliases: list[str] = []
    seen: set[str] = set()
    for raw in raw_values:
        alias = str(raw or "").strip()
        normalized = _normalize_product_alias(alias)
        if not alias or normalized in PRODUCT_ALIAS_STOPWORDS or len(normalized) < 2 or normalized in seen:
            continue
        seen.add(normalized)
        aliases.append(alias)
    return aliases


def spu_review_aliases(spu: ProductSPU) -> list[str]:
    info = loads(spu.extended_info_json, {}) if spu and spu.extended_info_json else {}
    aliases = []
    if isinstance(info, dict):
        review_aliases = info.get("review_aliases") or []
        aliases.extend(review_aliases if isinstance(review_aliases, list) else [review_aliases])
        pre_review = info.get("pre_review") or {}
        if isinstance(pre_review, dict):
            pre_review_aliases = pre_review.get("aliases") or []
            aliases.extend(pre_review_aliases if isinstance(pre_review_aliases, list) else [pre_review_aliases])
    return normalize_product_review_aliases(aliases)


def update_spu_review_aliases(session: Session, spu_uuid: str, aliases) -> ProductSPU:
    spu = (
        filter_finished_inventory_spus(session.query(ProductSPU))
        .filter(ProductSPU.id == spu_uuid)
        .one_or_none()
    )
    if spu is None:
        raise ValueError("只能为成品库存中的 SPU 维护预审别名")
    info = loads(spu.extended_info_json, {}) if spu.extended_info_json else {}
    if not isinstance(info, dict):
        info = {}
    info["review_aliases"] = normalize_product_review_aliases(aliases)
    spu.extended_info_json = json.dumps(info, ensure_ascii=False)
    spu.updated_at = now_utc()
    session.flush()
    return spu


def _spu_search_score(spu: ProductSPU, query: str) -> int:
    normalized_query = _normalize_product_alias(query)
    if not normalized_query:
        return 0
    fields = [
        ("spu_id", spu.spu_id, 100),
        ("name", spu.name, 90),
        ("alias", spu_review_aliases(spu), 95),
        ("brand", spu.brand, 60),
        ("category", spu.category, 50),
    ]
    best_score = 0
    for field_name, value, weight in fields:
        values = value if isinstance(value, list) else [value]
        for raw_value in values:
            normalized_value = _normalize_product_alias(raw_value)
            if not normalized_value:
                continue
            if normalized_value == normalized_query:
                score = weight + 30
            elif normalized_value.startswith(normalized_query):
                score = weight + 20
            elif normalized_query in normalized_value:
                score = weight + 10
            elif field_name in {"name", "alias"} and normalized_value in normalized_query:
                score = weight
            else:
                score = 0
            if score:
                best_score = max(best_score, score)
    return best_score


def get_spus(session: Session, skip: int = 0, limit: int = 100, query: str = None) -> tuple[list[ProductSPU], int]:
    q = filter_finished_inventory_spus(session.query(ProductSPU))
    query = str(query or "").strip()
    if query:
        scored_items = [
            (spu, _spu_search_score(spu, query))
            for spu in q.all()
        ]
        matched = [(spu, score) for spu, score in scored_items if score > 0]
        matched.sort(key=lambda item: (-item[1], item[0].spu_id or "", item[0].name or ""))
        total = len(matched)
        return [spu for spu, _ in matched[skip:skip + limit]], total
    total = q.count()
    items = q.order_by(ProductSPU.created_at.desc()).offset(skip).limit(limit).all()
    return items, total

def _sku_search_score(sku: ProductSKU, query: str) -> int:
    normalized_query = _normalize_product_alias(query)
    if not normalized_query:
        return 0
    spu = sku.spu
    # 从 attributes_json 提取 OMS 英文名称作为可搜索字段
    attrs = loads(sku.attributes_json, {}) if sku.attributes_json else {}
    oms_en_name = attrs.get("oms_en_name", "") if isinstance(attrs, dict) else ""
    fields = [
        ("sku_id", sku.sku_id, 110),
        ("spu_id", spu.spu_id if spu else "", 95),
        ("name", spu.name if spu else "", 90),
        ("model", sku.model or "", 85),
        ("alias", spu_review_aliases(spu) if spu else [], 95),
        ("en_name", oms_en_name, 80),
        ("brand", spu.brand if spu else "", 60),
        ("category", spu.category if spu else "", 50),
    ]
    best_score = 0
    for field_name, value, weight in fields:
        values = value if isinstance(value, list) else [value]
        for raw_value in values:
            normalized_value = _normalize_product_alias(raw_value)
            if not normalized_value:
                continue
            if normalized_value == normalized_query:
                score = weight + 30
            elif normalized_value.startswith(normalized_query):
                score = weight + 20
            elif normalized_query in normalized_value:
                score = weight + 10
            elif field_name in {"name", "alias"} and normalized_value in normalized_query:
                score = weight
            else:
                score = 0
            if score:
                best_score = max(best_score, score)
    return best_score


def get_skus(session: Session, skip: int = 0, limit: int = 100, spu_id: str = None, spu_uuid: str = None, query: str = None, crm_semantic: bool = False) -> tuple[list[ProductSKU], int]:
    query = str(query or "").strip()
    if crm_semantic and query:
        matched_skus = semantic_match_skus(session, query, limit=limit * 2)
        if matched_skus:
            # 过滤以满足 spu_uuid / spu_id 条件
            if spu_uuid:
                matched_skus = [sku for sku in matched_skus if sku.spu_uuid == spu_uuid]
            if spu_id:
                matched_skus = [sku for sku in matched_skus if sku.spu and sku.spu.spu_id == spu_id]
            total = len(matched_skus)
            return matched_skus[skip:skip + limit], total

    q = session.query(ProductSKU).join(ProductSPU, ProductSKU.spu_uuid == ProductSPU.id).filter(ProductSKU.status == "Active", ProductSPU.category == FINISHED_CATEGORY)
    if spu_uuid:
        q = q.where(ProductSKU.spu_uuid == spu_uuid)
    if spu_id:
        q = q.where(ProductSPU.spu_id == spu_id)
    if query:
        scored_items = [
            (sku, _sku_search_score(sku, query))
            for sku in q.all()
        ]
        matched = [(sku, score) for sku, score in scored_items if score > 0]
        matched.sort(key=lambda item: (-item[1], item[0].sku_id or "", item[0].spu.spu_id if item[0].spu else ""))
        total = len(matched)
        return [sku for sku, _ in matched[skip:skip + limit]], total
    total = q.count()
    items = q.order_by(ProductSKU.created_at.desc()).offset(skip).limit(limit).all()
    return items, total

def get_channel_pricing(session: Session, skip: int = 0, limit: int = 100, sku_id: str = None, sku_uuid: str = None, query: str = None) -> tuple[list[ChannelPricing], int]:
    q = (
        session.query(ChannelPricing)
        .join(ProductSKU, ProductSKU.id == ChannelPricing.sku_uuid)
    )
    q = filter_finished_inventory_skus(q)
    if sku_uuid:
        q = q.where(ChannelPricing.sku_uuid == sku_uuid)
    if sku_id:
        q = q.where(ProductSKU.sku_id == sku_id)
    query = str(query or "").strip()
    if query:
        scored_items = [
            (pricing, _sku_search_score(pricing.sku, query))
            for pricing in q.all()
        ]
        matched = [(pricing, score) for pricing, score in scored_items if score > 0]
        matched.sort(key=lambda item: (-item[1], item[0].sku.sku_id if item[0].sku else "", item[0].channel or ""))
        total = len(matched)
        return [pricing for pricing, _ in matched[skip:skip + limit]], total
    total = q.count()
    items = q.order_by(ChannelPricing.updated_at.desc()).offset(skip).limit(limit).all()
    return items, total


def product_review_readiness(session: Session, *, channel: str = "default", limit: int = 20) -> dict:
    channel = str(channel or "default").strip() or "default"
    active_skus = (
        filter_finished_inventory_skus(session.query(ProductSKU))
        .filter(ProductSKU.status == "Active")
        .order_by(ProductSKU.sku_id)
        .all()
    )
    sku_ids = [sku.id for sku in active_skus]
    pricing_by_sku = {
        pricing.sku_uuid: pricing
        for pricing in session.query(ChannelPricing).filter(
            ChannelPricing.channel == channel,
            ChannelPricing.sku_uuid.in_(sku_ids) if sku_ids else False,
        ).all()
    }
    issues: list[dict] = []
    spu_seen: dict[str, ProductSPU] = {}
    alias_owners: dict[str, list[ProductSPU]] = {}
    missing_price_count = 0
    incomplete_price_count = 0
    for sku in active_skus:
        spu = sku.spu
        if spu:
            spu_seen[spu.id] = spu
            for alias in spu_review_aliases(spu):
                alias_owners.setdefault(_normalize_product_alias(alias), []).append(spu)
        pricing = pricing_by_sku.get(sku.id)
        base_min_price = pricing.map_price if pricing and pricing.map_price is not None else (pricing.tier_a_price if pricing else None)
        item_base = {
            "sku_uuid": sku.id,
            "sku_id": sku.sku_id,
            "spu_uuid": sku.spu_uuid,
            "spu_id": spu.spu_id if spu else "",
            "product_name": spu.name if spu else "",
            "channel": channel,
        }
        if pricing is None:
            missing_price_count += 1
            issues.append({
                **item_base,
                "severity": "blocker",
                "issue_type": "missing_price",
                "message": f"渠道 {channel} 未配置价格规则",
                "action": "configure_pricing",
            })
        elif base_min_price is None:
            incomplete_price_count += 1
            issues.append({
                **item_base,
                "severity": "blocker",
                "issue_type": "incomplete_price",
                "message": f"渠道 {channel} 价格规则缺少 MAP 或 A 档最低价",
                "action": "configure_pricing",
            })
    no_manual_alias_count = 0
    for spu in sorted(spu_seen.values(), key=lambda row: row.spu_id):
        if spu_review_aliases(spu):
            continue
        no_manual_alias_count += 1
        issues.append({
            "severity": "warning",
            "issue_type": "missing_alias",
            "message": "未维护人工预审别名，订单只写业务叫法时可能漏匹配",
            "action": "configure_alias",
            "spu_uuid": spu.id,
            "spu_id": spu.spu_id,
            "product_name": spu.name,
            "review_aliases": [],
        })
    duplicate_alias_count = 0
    for normalized_alias, owners in alias_owners.items():
        unique_owners = {owner.id: owner for owner in owners}
        if not normalized_alias or len(unique_owners) <= 1:
            continue
        duplicate_alias_count += 1
        owner_list = list(unique_owners.values())
        issues.append({
            "severity": "blocker",
            "issue_type": "duplicate_alias",
            "message": f"预审别名重复：{', '.join(owner.spu_id for owner in owner_list[:4])}",
            "action": "review_alias",
            "alias": owner_list[0] and next((alias for alias in spu_review_aliases(owner_list[0]) if _normalize_product_alias(alias) == normalized_alias), normalized_alias),
            "spu_uuid": owner_list[0].id,
            "spu_id": owner_list[0].spu_id,
            "product_name": owner_list[0].name,
            "owners": [{"spu_uuid": owner.id, "spu_id": owner.spu_id, "name": owner.name} for owner in owner_list],
        })
    active_promotions = (
        session.query(PromotionRule)
        .filter(
            PromotionRule.is_active == True,
            (PromotionRule.channel == channel) | (PromotionRule.channel == None),
        )
        .order_by(PromotionRule.priority.desc(), PromotionRule.created_at.desc())
        .all()
    )
    invalid_promotion_count = 0
    finished_sku_ids = set(sku_ids)
    promotion_owners: dict[tuple[str, str], list[PromotionRule]] = {}
    for promotion in active_promotions:
        if promotion.sku_uuid in finished_sku_ids:
            promotion_owners.setdefault((_normalize_product_alias(promotion.name), promotion.sku_uuid), []).append(promotion)
            continue
        invalid_promotion_count += 1
        reason = "未绑定成品 SKU" if not promotion.sku_uuid else "绑定的 SKU 不属于成品库存"
        issues.append({
            "severity": "blocker",
            "issue_type": "invalid_promotion",
            "message": f"促销规则「{promotion.name}」{reason}，不会参与订单预审",
            "action": "configure_promotion",
            "promotion_id": promotion.id,
            "promotion_name": promotion.name,
            "channel": promotion.channel or "通用",
        })
    duplicate_promotion_count = 0
    for (normalized_name, sku_uuid), promotions in promotion_owners.items():
        if not normalized_name or len(promotions) <= 1:
            continue
        duplicate_promotion_count += 1
        sku = promotions[0].sku
        spu = sku.spu if sku else None
        issues.append({
            "severity": "blocker",
            "issue_type": "duplicate_promotion",
            "message": f"SKU {sku.sku_id if sku else sku_uuid} 的促销规则「{promotions[0].name}」重复配置，订单预审无法判断应使用哪一条",
            "action": "configure_promotion",
            "promotion_id": promotions[0].id,
            "promotion_name": promotions[0].name,
            "sku_uuid": sku_uuid,
            "sku_id": sku.sku_id if sku else "",
            "spu_id": spu.spu_id if spu else "",
            "product_name": spu.name if spu else "",
            "channel": channel,
        })
    severity_rank = {"blocker": 0, "warning": 1}
    issues.sort(key=lambda item: (severity_rank.get(item["severity"], 2), item["issue_type"], item.get("spu_id") or item.get("sku_id") or ""))
    blocker_count = missing_price_count + incomplete_price_count + duplicate_alias_count + invalid_promotion_count + duplicate_promotion_count
    warning_count = no_manual_alias_count
    score = max(0, 100 - blocker_count * 5 - warning_count)
    return {
        "ok": True,
        "channel": channel,
        "summary": {
            "score": score,
            "finished_sku_count": len(active_skus),
            "finished_spu_count": len(spu_seen),
            "priced_sku_count": len(pricing_by_sku),
            "missing_price_count": missing_price_count,
            "incomplete_price_count": incomplete_price_count,
            "missing_alias_count": no_manual_alias_count,
            "duplicate_alias_count": duplicate_alias_count,
            "invalid_promotion_count": invalid_promotion_count,
            "duplicate_promotion_count": duplicate_promotion_count,
            "blocker_count": blocker_count,
            "warning_count": warning_count,
            "issue_count": len(issues),
        },
        "issues": issues[:limit],
        "total_issues": len(issues),
    }


def promotion_rule_binding_info(session: Session, rule: PromotionRule) -> dict:
    if not rule.sku_uuid:
        return {"status": "unbound", "label": "未绑定成品 SKU", "is_valid": False}
    sku = rule.sku
    if not sku or not sku.spu:
        return {"status": "missing_sku", "label": "绑定 SKU 不存在", "is_valid": False}
    has_finished_inventory = (
        session.query(ProductInventorySnapshot.id)
        .filter(ProductInventorySnapshot.material_code == sku.spu.spu_id)
        .first()
        is not None
    )
    if sku.spu.category != FINISHED_CATEGORY or not has_finished_inventory:
        return {"status": "non_finished_sku", "label": "绑定 SKU 不属于成品库存", "is_valid": False}
    return {"status": "ok", "label": "已绑定成品 SKU", "is_valid": True}


def _promotion_search_score(session: Session, rule: PromotionRule, query: str) -> int:
    normalized_query = _normalize_product_alias(query)
    if not normalized_query:
        return 0
    binding = promotion_rule_binding_info(session, rule)
    sku = rule.sku
    spu = sku.spu if sku else None
    fields = [
        (rule.name, 110),
        (sku.sku_id if sku else "", 100),
        (spu.spu_id if spu else "", 90),
        (spu.name if spu else "", 85),
        (rule.channel or "通用", 60),
        (binding["label"], 80),
        (binding["status"], 70),
    ]
    best_score = 0
    for raw_value, weight in fields:
        normalized_value = _normalize_product_alias(raw_value)
        if not normalized_value:
            continue
        if normalized_value == normalized_query:
            score = weight + 30
        elif normalized_value.startswith(normalized_query):
            score = weight + 20
        elif normalized_query in normalized_value:
            score = weight + 10
        elif normalized_value in normalized_query:
            score = weight
        else:
            score = 0
        best_score = max(best_score, score)
    return best_score


def get_promotions(session: Session, skip: int = 0, limit: int = 100, query: str = None) -> tuple[list[PromotionRule], int]:
    q = session.query(PromotionRule)
    query = str(query or "").strip()
    if query:
        scored_items = [(rule, _promotion_search_score(session, rule, query)) for rule in q.all()]
        matched = [(rule, score) for rule, score in scored_items if score > 0]
        matched.sort(key=lambda item: (-item[1], -(item[0].priority or 0), item[0].name or ""))
        total = len(matched)
        return [rule for rule, _ in matched[skip:skip + limit]], total
    total = q.count()
    items = q.order_by(PromotionRule.priority.desc(), PromotionRule.created_at.desc()).offset(skip).limit(limit).all()
    return items, total


def _validate_time_range(start_time: datetime | None, end_time: datetime | None, label: str) -> None:
    if start_time and end_time and start_time > end_time:
        raise ValueError(f"{label}开始时间不能晚于结束时间")


def _validate_nonnegative_price(label: str, value: int | None) -> None:
    if value is None:
        return
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{label}必须是有效金额")
    if number < 0:
        raise ValueError(f"{label}不能为负数")


def validate_channel_pricing_values(
    *,
    tier_a_price: int | None = None,
    tier_b_price: int | None = None,
    tier_c_price: int | None = None,
    map_price: int | None = None,
    max_price: int | None = None,
    promo_start_time: datetime | None = None,
    promo_end_time: datetime | None = None,
) -> None:
    prices = {
        "A档价格": tier_a_price,
        "B档价格": tier_b_price,
        "C档价格": tier_c_price,
        "底价(MAP)": map_price,
        "最高限价": max_price,
    }
    for label, value in prices.items():
        _validate_nonnegative_price(label, value)
    lower_prices = [value for value in [tier_a_price, tier_b_price, tier_c_price, map_price] if value is not None]
    if max_price is not None and lower_prices and max(lower_prices) > max_price:
        raise ValueError("最高限价不能低于已配置的最低价格")
    _validate_time_range(promo_start_time, promo_end_time, "渠道价格促销")


def validate_promotion_rule_values(
    *,
    name: str,
    discount_type: str,
    discount_value: int,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
) -> str:
    if not str(name or "").strip():
        raise ValueError("促销规则名称不能为空")
    normalized_type = str(discount_type or "").lower()
    if normalized_type not in {"percentage", "fixed_amount"}:
        raise ValueError("促销优惠类型无效")
    try:
        value = int(discount_value)
    except (TypeError, ValueError):
        raise ValueError("促销优惠数值必须是有效数字")
    if normalized_type == "percentage" and not 1 <= value <= 100:
        raise ValueError("比例折扣请输入 1-100 之间的数字")
    if normalized_type == "fixed_amount" and value <= 0:
        raise ValueError("固定减免金额必须大于 0")
    _validate_time_range(start_time, end_time, "促销规则")
    return normalized_type


def create_promotion_rule(
    session: Session,
    name: str,
    sku_uuid: str,
    discount_type: str,
    discount_value: int,
    channel: str | None = None,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
    priority: int = 0
) -> PromotionRule:
    ensure_finished_inventory_sku(session, sku_uuid)
    discount_type = validate_promotion_rule_values(
        name=name,
        discount_type=discount_type,
        discount_value=discount_value,
        start_time=start_time,
        end_time=end_time,
    )
    rule = PromotionRule(
        sku_uuid=sku_uuid,
        name=name,
        channel=channel,
        start_time=start_time,
        end_time=end_time,
        priority=priority,
        discount_type=discount_type,
        discount_value=discount_value
    )
    session.add(rule)
    session.flush() # Ensure ID is generated
    return rule


def update_promotion_rule(
    session: Session,
    rule_id: str,
    **kwargs
) -> PromotionRule | None:
    from sqlalchemy import select
    rule = session.execute(select(PromotionRule).filter(PromotionRule.id == rule_id)).scalar_one_or_none()
    if not rule:
        return None
    merged = {
        "name": rule.name,
        "discount_type": rule.discount_type,
        "discount_value": rule.discount_value,
        "start_time": rule.start_time,
        "end_time": rule.end_time,
        **kwargs,
    }
    if "sku_uuid" in kwargs and not kwargs.get("sku_uuid"):
        raise ValueError("促销规则必须绑定成品 SKU")
    if kwargs.get("sku_uuid"):
        ensure_finished_inventory_sku(session, kwargs["sku_uuid"])
    merged["discount_type"] = validate_promotion_rule_values(
        name=merged.get("name"),
        discount_type=merged.get("discount_type"),
        discount_value=merged.get("discount_value"),
        start_time=merged.get("start_time"),
        end_time=merged.get("end_time"),
    )
    kwargs["discount_type"] = merged["discount_type"]
    
    for k, v in kwargs.items():
        if hasattr(rule, k):
            setattr(rule, k, v)
    
    rule.updated_at = now_utc()
    session.flush()
    return rule


def delete_promotion_rule(
    session: Session,
    rule_id: str
) -> bool:
    from sqlalchemy import select
    rule = session.execute(select(PromotionRule).filter(PromotionRule.id == rule_id)).scalar_one_or_none()
    if not rule:
        return False
    session.delete(rule)
    return True


def toggle_promotion_rule(
    session: Session,
    rule_id: str,
    is_active: bool
) -> PromotionRule | None:
    from sqlalchemy import select
    rule = session.execute(select(PromotionRule).filter(PromotionRule.id == rule_id)).scalar_one_or_none()
    if not rule:
        return None
    rule.is_active = is_active
    rule.updated_at = now_utc()
    session.flush()
    return rule

def create_spu(session: Session, spu_id: str, name: str, brand: str = None, category: str = None) -> ProductSPU:
    spu = ProductSPU(spu_id=spu_id, name=name, brand=brand, category=category)
    session.add(spu)
    return spu

def create_sku(session: Session, spu_uuid: str, sku_id: str, attributes: dict = None) -> ProductSKU:
    spu = (
        filter_finished_inventory_spus(session.query(ProductSPU))
        .filter(ProductSPU.id == spu_uuid)
        .one_or_none()
    )
    if spu is None:
        raise ValueError("SKU 只能绑定到成品库存中的 SPU")
    sku = ProductSKU(spu_uuid=spu_uuid, sku_id=sku_id, attributes_json=json.dumps(attributes or {}))
    session.add(sku)
    return sku

def set_channel_pricing(
    session: Session, 
    sku_uuid: str, 
    channel: str, 
    tier_a_price: int = None, 
    tier_b_price: int = None, 
    tier_c_price: int = None, 
    map_price: int = None, 
    promo_start_time: datetime = None,
    promo_end_time: datetime = None,
    currency: str = "USD"
) -> ChannelPricing:
    from sqlalchemy import select
    ensure_finished_inventory_sku(session, sku_uuid)
    validate_channel_pricing_values(
        tier_a_price=tier_a_price,
        tier_b_price=tier_b_price,
        tier_c_price=tier_c_price,
        map_price=map_price,
        promo_start_time=promo_start_time,
        promo_end_time=promo_end_time,
    )
    pricing = session.execute(select(ChannelPricing).where(ChannelPricing.sku_uuid == sku_uuid, ChannelPricing.channel == channel)).scalars().first()
    if not pricing:
        pricing = ChannelPricing(sku_uuid=sku_uuid, channel=channel)
        session.add(pricing)
    
    pricing.tier_a_price = tier_a_price
    pricing.tier_b_price = tier_b_price
    pricing.tier_c_price = tier_c_price
    pricing.map_price = map_price
    pricing.promo_start_time = promo_start_time
    pricing.promo_end_time = promo_end_time
    pricing.currency = currency
    pricing.updated_at = now_utc()
    return pricing


def config_bool(session: Session, key: str, default: bool) -> bool:
    row = session.get(SystemConfig, key)
    if row is None:
        return default
    value = str(row.value or "").strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    return default


def parse_price_to_cents(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value * 100))
    text = str(value).strip()
    if not text:
        return None
    match = re.search(
        r"(?:(?P<prefix>[¥￥$]|rmb|cny|usd|us\$|人民币|美元|美金)\s*)?"
        r"(?P<number>\d+(?:,\d{3})*(?:\.\d+)?|\d+)"
        r"(?:\s*(?P<unit>元|块|rmb|cny|usd|us\$|人民币|美元|美金|分))?"
        r"(?:\s*/\s*(?:台|套|件|个|pcs?|piece|unit))?",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    number = float(match.group("number").replace(",", ""))
    unit = (match.group("unit") or "").lower()
    if unit == "分":
        return int(round(number))
    return int(round(number * 100))


PRICE_PATTERN = re.compile(
    r"(?:单价|价格|售价|报价|含税价|成交价|销售价|unit\s*price|price)\s*[:：]?\s*"
    r"(?P<price>(?:[¥￥$]|rmb|cny|usd|us\$|人民币|美元|美金)?\s*"
    r"\d+(?:,\d{3})*(?:\.\d+)?"
    r"(?:\s*(?:元|块|rmb|cny|usd|us\$|人民币|美元|美金|分))?"
    r"(?:\s*/\s*(?:台|套|件|个|pcs?|piece|unit))?)",
    flags=re.IGNORECASE,
)

PRODUCT_ALIAS_STOPWORDS = {
    "成品",
    "主机",
    "套装",
    "三维扫描仪",
    "扫描仪",
    "标准版",
    "国内版",
    "海外版",
    "中文版",
    "英文版",
    "美规",
    "欧规",
    "英规",
    "澳规",
    "4规",
}


def _normalize_product_alias(value: str | None) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[（(]\s*4\s*规\s*[）)]", "4规", text)
    text = re.sub(r"[\s_\-/,，、:：;；|]+", "", text)
    text = re.sub(r"[()（）\[\]【】\"'“”‘’]", "", text)
    return text


def _product_name_match_tokens(value: str | None) -> list[str]:
    text = str(value or "").strip().lower()
    tokens: list[str] = []
    seen: set[str] = set()
    for token in re.findall(r"[a-z][a-z0-9+]*", text):
        normalized = _normalize_product_alias(token)
        if len(normalized) < 3 or normalized in PRODUCT_ALIAS_STOPWORDS or normalized in seen:
            continue
        seen.add(normalized)
        tokens.append(normalized)
    return tokens


def _strong_alias(value: str | None) -> str | None:
    alias = str(value or "").strip()
    normalized = _normalize_product_alias(alias)
    if not normalized or normalized in PRODUCT_ALIAS_STOPWORDS:
        return None
    has_chinese = bool(re.search(r"[\u4e00-\u9fff]", normalized))
    has_digit = bool(re.search(r"\d", normalized))
    has_alpha = bool(re.search(r"[a-z]", normalized))
    if has_chinese and len(normalized) >= 4:
        return alias
    if has_alpha and has_digit and len(normalized) >= 3:
        return alias
    if has_alpha and len(normalized) >= 5:
        return alias
    return None


def _derived_product_name_aliases(value: str | None) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    variants = [
        text,
        re.sub(r"[（(][^）)]*[）)]", "", text).strip(),
        re.sub(r"^\s*(?:三维扫描仪|扫描仪)\s*", "", text).strip(),
        re.sub(r"^\s*(?:三维扫描仪|扫描仪)\s*", "", re.sub(r"[（(][^）)]*[）)]", "", text)).strip(),
    ]
    aliases: list[str] = []
    seen: set[str] = set()
    for variant in variants:
        alias = _strong_alias(variant)
        normalized = _normalize_product_alias(alias)
        if not alias or normalized in seen:
            continue
        seen.add(normalized)
        aliases.append(alias)
    return aliases


def _sku_match_aliases(sku: ProductSKU) -> list[str]:
    spu = sku.spu
    attrs = loads(sku.attributes_json, {}) if sku.attributes_json else {}
    extended = loads(spu.extended_info_json, {}) if spu and spu.extended_info_json else {}
    erp = extended.get("erp", {}) if isinstance(extended, dict) else {}
    candidates = [
        sku.sku_id,
        sku.model,
        sku.version,
        attrs.get("erp_material_name"),
        attrs.get("erp_specification"),
        spu.spu_id if spu else None,
        spu.name if spu else None,
        spu.name_en if spu else None,
        erp.get("material_number") if isinstance(erp, dict) else None,
        erp.get("specification") if isinstance(erp, dict) else None,
    ]
    aliases: list[str] = []
    seen: set[str] = set()
    for alias in spu_review_aliases(spu) if spu else []:
        normalized = _normalize_product_alias(alias)
        if normalized and normalized not in seen:
            seen.add(normalized)
            aliases.append(alias)
    for candidate in candidates:
        for alias in [*_derived_product_name_aliases(candidate)]:
            normalized = _normalize_product_alias(alias)
            if not alias or normalized in seen:
                continue
            seen.add(normalized)
            aliases.append(alias)
    aliases.sort(key=lambda item: len(_normalize_product_alias(item)), reverse=True)
    return aliases


def _find_product_alias_index(text: str, alias: str) -> int:
    normalized_text = _normalize_product_alias(text)
    normalized_alias = _normalize_product_alias(alias)
    if not normalized_alias:
        return -1
    return normalized_text.find(normalized_alias)


def _extract_price_near_alias(text: str, alias: str) -> int | None:
    index = _find_product_alias_index(text, alias)
    if index < 0:
        return _extract_price_near_sku(text, alias)
    normalized_text = _normalize_product_alias(text)
    # Use the normalized offset only as a rough anchor, then keep a wide window in the original text.
    original_anchor = min(len(text), max(0, index))
    windows = [text[max(0, original_anchor - 120): original_anchor + len(alias) + 180], text]
    for window in windows:
        match = PRICE_PATTERN.search(window)
        if match:
            return parse_price_to_cents(match.group("price"))
    if normalized_text:
        return None
    return _extract_price_near_sku(text, alias)


def product_review_alias_candidate_from_text(text: str) -> str:
    source = str(text or "").strip()
    if not source:
        return ""
    line = next((part.strip() for part in re.split(r"[\r\n]+", source) if part.strip()), source)
    line = re.sub(r"^\s*(?:客户下单|客户需要|产品|物料|物料名称|名称|型号|品名)\s*[:：]\s*", "", line)
    line = re.split(r"(?:数量|单价|价格|售价|报价|期望交期|订单号|收件|地址)\s*[:：]?", line, maxsplit=1)[0]
    line = re.sub(r"^[：:，,、.\s]+|[：:，,、.\s]+$", "", line)
    if len(line) > 80:
        line = line[:80].strip()
    return line


def _product_review_search_tokens(text: str) -> list[str]:
    normalized = _normalize_product_alias(text)
    raw_tokens = re.findall(r"[a-z]+\d*[a-z\d]*|\d+[a-z]+[a-z\d]*|[\u4e00-\u9fff]{2,}", normalized)
    tokens: list[str] = []
    seen: set[str] = set()
    for token in raw_tokens:
        if token in PRODUCT_ALIAS_STOPWORDS or len(token) < 2 or token in seen:
            continue
        seen.add(token)
        tokens.append(token)
    return tokens


def suggest_product_review_candidates(session: Session, text: str, *, limit: int = 5) -> list[dict]:
    source = str(text or "")
    alias_candidate = product_review_alias_candidate_from_text(source)
    query_norm = _normalize_product_alias(alias_candidate or source)
    tokens = _product_review_search_tokens(f"{alias_candidate} {source}")
    if not query_norm and not tokens:
        return []
    active_skus = (
        session.query(ProductSKU)
        .join(ProductSPU, ProductSKU.spu_uuid == ProductSPU.id)
        .filter(ProductSKU.status == "Active", ProductSPU.status == "Active")
        .all()
    )
    best_by_spu: dict[str, dict] = {}
    for sku in active_skus:
        spu = sku.spu
        if not spu:
            continue
        best_score = 0
        best_alias = ""
        for alias in _sku_match_aliases(sku):
            alias_norm = _normalize_product_alias(alias)
            if not alias_norm:
                continue
            score = 0
            if query_norm and (alias_norm in query_norm or query_norm in alias_norm):
                score = 100 + min(len(alias_norm), len(query_norm))
            for token in tokens:
                if token and (token in alias_norm or alias_norm in token):
                    score = max(score, 40 + min(len(token), len(alias_norm)))
                elif len(token) >= 4 and len(alias_norm) >= 4:
                    ratio = SequenceMatcher(None, token, alias_norm).ratio()
                    if ratio >= 0.66:
                        score = max(score, int(ratio * 35))
            if score > best_score:
                best_score = score
                best_alias = alias
        if best_score <= 0:
            continue
        current = best_by_spu.get(spu.id)
        if current is None or best_score > current["score"]:
            best_by_spu[spu.id] = {
                "id": spu.id,
                "spu_id": spu.spu_id,
                "name": spu.name,
                "category": spu.category,
                "sku_id": sku.sku_id,
                "matched_alias": best_alias,
                "suggested_alias": alias_candidate,
                "review_aliases": spu_review_aliases(spu),
                "score": best_score,
            }
    return sorted(best_by_spu.values(), key=lambda item: item["score"], reverse=True)[:limit]


def product_name_sku_candidates(session: Session, query: str, *, limit: int = 5) -> list[dict]:
    source = str(query or "").strip()
    query_norm = _normalize_product_alias(source)
    query_tokens = _product_name_match_tokens(source)
    if not query_norm:
        return []
    active_skus = (
        session.query(ProductSKU)
        .join(ProductSPU, ProductSKU.spu_uuid == ProductSPU.id)
        .filter(ProductSKU.status == "Active", ProductSPU.status == "Active")
        .all()
    )
    candidates: list[dict] = []
    for sku in active_skus:
        spu = sku.spu
        if not spu:
            continue
        fields: list[tuple[str, str, int]] = []
        fields.extend((alias, "alias", 300) for alias in spu_review_aliases(spu))
        fields.append((spu.name, "name", 200))
        fields.append((spu.name_en, "english_name", 100))
        attrs = loads(sku.attributes_json, {}) if sku.attributes_json else {}
        fields.append((attrs.get("english_name") or attrs.get("name_en"), "english_name", 100))
        best: dict | None = None
        for raw_value, source_name, priority in fields:
            value = str(raw_value or "").strip()
            value_norm = _normalize_product_alias(value)
            if not value_norm:
                continue
            score = 0
            confidence = 0
            source_base = {"alias": 96, "name": 90, "english_name": 84}.get(source_name, 80)
            if query_norm == value_norm:
                score = priority + 100
                confidence = source_base + 3
            elif value_norm in query_norm:
                score = priority + min(90, len(value_norm) * 4)
                confidence = source_base
            elif query_norm in value_norm:
                score = priority + min(80, len(query_norm) * 3)
                confidence = max(75, source_base - 4)
            elif query_tokens and any(token in value_norm for token in query_tokens):
                token = next(token for token in query_tokens if token in value_norm)
                score = priority + min(70, len(token) * 8)
                confidence = max(82, source_base - 6)
            else:
                ratio = SequenceMatcher(None, query_norm, value_norm).ratio()
                if ratio >= 0.72:
                    score = priority + int(ratio * 70)
                    confidence = min(source_base - 8, int(ratio * 100))
            if score <= 0:
                continue
            candidate = {
                "sku_id": sku.sku_id,
                "spu_id": spu.spu_id,
                "product_name": spu.name,
                "matched_value": value,
                "match_source": source_name,
                "confidence": max(0, min(99, confidence)),
                "score": score,
            }
            if best is None or (candidate["score"], candidate["confidence"]) > (best["score"], best["confidence"]):
                best = candidate
        if best:
            candidates.append(best)
    candidates.sort(key=lambda item: (item["score"], item["confidence"]), reverse=True)
    return candidates[:limit]


def match_sku_by_product_name(session: Session, query: str, *, min_confidence: int = 80) -> dict:
    candidates = product_name_sku_candidates(session, query, limit=5)
    if not candidates:
        return {"matched": False, "reason": "not_found", "candidates": []}
    best = candidates[0]
    if int(best.get("confidence") or 0) < min_confidence:
        return {"matched": False, "reason": "low_confidence", "candidates": candidates}
    ties = [item for item in candidates if item.get("confidence") == best.get("confidence") and item.get("sku_id") != best.get("sku_id")]
    if ties:
        return {"matched": False, "reason": "ambiguous", "candidates": candidates}
    return {"matched": True, "sku_id": best["sku_id"], "confidence": best["confidence"], "match_source": best["match_source"], "matched_value": best["matched_value"], "candidates": candidates}


def extract_order_products_from_text(session: Session, text: str, *, channel: str = "default") -> list[dict]:
    source = str(text or "")
    if not source.strip():
        return []
    active_skus = (
        filter_finished_inventory_skus(session.query(ProductSKU))
        .filter(ProductSKU.status == "Active")
        .order_by(ProductSKU.sku_id.desc())
        .all()
    )
    if not active_skus:
        return []
    active_promotion_query = (
        session.query(PromotionRule)
        .join(ProductSKU, ProductSKU.id == PromotionRule.sku_uuid)
    )
    active_promotions = (
        filter_finished_inventory_skus(active_promotion_query)
        .filter(
            PromotionRule.is_active == True,
            (PromotionRule.channel == channel) | (PromotionRule.channel == None),
        )
        .all()
    )
    promotion_names = [promo.name for promo in active_promotions if promo.name and promo.name in source]
    items: list[dict] = []
    lower_source = source.lower()
    matched_sku_ids: set[str] = set()
    matched_alias_spans: list[tuple[int, int]] = []
    sku_aliases = {sku.sku_id: _sku_match_aliases(sku) for sku in active_skus if sku.sku_id}
    alias_counts: dict[str, int] = {}
    for aliases in sku_aliases.values():
        for alias in aliases:
            normalized_alias = _normalize_product_alias(alias)
            alias_counts[normalized_alias] = alias_counts.get(normalized_alias, 0) + 1
    for sku in sorted(active_skus, key=lambda row: len(row.sku_id or ""), reverse=True):
        sku_id = str(sku.sku_id or "").strip()
        if not sku_id or sku_id.lower() not in lower_source:
            continue
        unit_price = _extract_price_near_sku(source, sku_id)
        matched_sku_ids.add(sku_id)
        sku_index = _normalize_product_alias(source).find(_normalize_product_alias(sku_id))
        if sku_index >= 0:
            matched_alias_spans.append((sku_index, sku_index + len(_normalize_product_alias(sku_id))))
        items.append({
            "sku_id": sku_id,
            "sku_code": sku_id,
            "unit_price": unit_price,
            "promotion_applied": promotion_names,
            "match_source": "sku_code",
            "match_alias": sku_id,
        })
    alias_matches: list[tuple[int, int, ProductSKU, str]] = []
    for sku in active_skus:
        sku_id = str(sku.sku_id or "").strip()
        if not sku_id or sku_id in matched_sku_ids:
            continue
        for alias in sku_aliases.get(sku_id, []):
            normalized_alias = _normalize_product_alias(alias)
            if normalized_alias != _normalize_product_alias(sku_id) and alias_counts.get(normalized_alias, 0) > 1:
                continue
            alias_index = _find_product_alias_index(source, alias)
            if alias_index < 0:
                continue
            alias_matches.append((alias_index, alias_index + len(normalized_alias), sku, alias))
    alias_matches.sort(key=lambda item: (item[1] - item[0]), reverse=True)
    for start, end, sku, alias in alias_matches:
        sku_id = str(sku.sku_id or "").strip()
        if not sku_id or sku_id in matched_sku_ids:
            continue
        if any(start < used_end and end > used_start for used_start, used_end in matched_alias_spans):
            continue
        unit_price = _extract_price_near_alias(source, alias)
        matched_sku_ids.add(sku_id)
        matched_alias_spans.append((start, end))
        items.append({
            "sku_id": sku_id,
            "sku_code": sku_id,
            "unit_price": unit_price,
            "promotion_applied": promotion_names,
            "match_source": "product_alias",
            "match_alias": alias,
        })
    return items


def _extract_price_near_sku(text: str, sku_id: str) -> int | None:
    sku_index = text.lower().find(sku_id.lower())
    windows: list[str] = []
    if sku_index >= 0:
        windows.append(text[max(0, sku_index - 80): sku_index + len(sku_id) + 120])
    windows.append(text)
    for window in windows:
        match = PRICE_PATTERN.search(window)
        if match:
            return parse_price_to_cents(match.group("price"))
    return None


def normalize_extracted_product_item(item: dict) -> dict:
    normalized = dict(item)
    sku_id = (
        normalized.get("sku_id")
        or normalized.get("sku_code")
        or normalized.get("product_code")
        or normalized.get("code")
        or normalized.get("product_id")
    )
    normalized["sku_id"] = str(sku_id).strip() if sku_id is not None else ""
    normalized["sku_code"] = normalized.get("sku_code") or normalized["sku_id"]
    normalized["unit_price"] = parse_price_to_cents(normalized.get("unit_price"))
    promotions = normalized.get("promotion_applied") or normalized.get("promotions") or []
    if isinstance(promotions, str):
        promotions = [promotions]
    normalized["promotion_applied"] = [str(item).strip() for item in promotions if str(item).strip()]
    return normalized


def promotion_discounted_min_price(base_min_price: int | None, promotion: PromotionRule) -> int | None:
    if base_min_price is None:
        return None
    discount_type = str(promotion.discount_type or "").lower()
    discount_value = int(promotion.discount_value or 0)
    if discount_value <= 0:
        return base_min_price
    if discount_type == "percentage":
        discount = min(discount_value, 100)
        return max(0, int(round(base_min_price * (100 - discount) / 100)))
    if discount_type == "fixed_amount":
        return max(0, base_min_price - discount_value)
    return base_min_price


def extract_order_products_for_review(session: Session, product_summary: str, source_text: str = "", *, channel: str = "default") -> list[dict]:
    local_items = extract_order_products_from_text(session, f"{product_summary or ''}\n{source_text or ''}", channel=channel)
    if local_items:
        return local_items
    if not config_bool(session, "product_price_review_llm_enabled", False):
        return []
    return extract_products_with_llm(session, product_summary or source_text)

def review_order_products(session: Session, extracted_items: list[dict], channel: str = "default") -> list[dict]:
    """
    Review a list of extracted material items against the database.
    Each item in extracted_items is expected to have:
      - 'sku_id'
      - 'unit_price' (integer representing cents/base unit)
      - 'promotion_applied' (optional list of promotion names)

    Returns the same list with added 'review_result' and 'risk_flags' keys.
    """
    results = []
    now = now_utc()
    
    # Get active promotions for the channel
    active_promotion_query = (
        session.query(PromotionRule)
        .join(ProductSKU, ProductSKU.id == PromotionRule.sku_uuid)
    )
    active_promotions = filter_finished_inventory_skus(active_promotion_query).filter(
        PromotionRule.is_active == True,
        (PromotionRule.channel == channel) | (PromotionRule.channel == None),
        (PromotionRule.start_time == None) | (PromotionRule.start_time <= now),
        (PromotionRule.end_time == None) | (PromotionRule.end_time >= now)
    ).all()
    active_promo_names: dict[str, list[PromotionRule]] = {}
    for promotion in active_promotions:
        active_promo_names.setdefault(promotion.name, []).append(promotion)

    require_unit_price = config_bool(session, "product_price_review_require_unit_price", False)
    for raw_item in extracted_items:
        item = normalize_extracted_product_item(raw_item)
        review = {"status": "Pass", "risk_flags": []}
        
        sku_id = item.get("sku_id")
        unit_price = item.get("unit_price")
        
        if not sku_id:
            review["status"] = "Warning"
            review["risk_flags"].append("未识别到 SKU，无法完成价格审查")
            item["review"] = review
            results.append(item)
            continue
            
        sku = (
            filter_finished_inventory_skus(session.query(ProductSKU))
            .filter(ProductSKU.sku_id == sku_id, ProductSKU.status == "Active")
            .one_or_none()
        )
        if not sku:
            review["status"] = "Exception"
            review["risk_flags"].append(f"SKU 不存在、未启用，或不属于成品库存：{sku_id}")
            item["review"] = review
            results.append(item)
            continue

        # Check Pricing
        pricing = session.execute(select(ChannelPricing).where(ChannelPricing.sku_uuid == sku.id, ChannelPricing.channel == channel)).scalars().first()
        item.update({
            "sku_uuid": sku.id,
            "spu_uuid": sku.spu_uuid,
            "spu_id": sku.spu.spu_id if sku.spu else "",
            "product_name": sku.spu.name if sku.spu else "",
            "pricing_configured": pricing is not None,
        })
        if unit_price is not None:
            if pricing:
                # Determine applicable minimum price based on time
                is_promo_period = False
                if pricing.promo_start_time and pricing.promo_end_time:
                    if pricing.promo_start_time <= now <= pricing.promo_end_time:
                        is_promo_period = True
                elif pricing.promo_start_time and now >= pricing.promo_start_time:
                    is_promo_period = True
                elif pricing.promo_end_time and now <= pricing.promo_end_time:
                    is_promo_period = True
                
                # Base minimum price
                base_min_price = pricing.map_price if pricing.map_price is not None else pricing.tier_a_price
                effective_min_price = base_min_price
                if is_promo_period:
                    candidates = [x for x in [pricing.tier_b_price, pricing.tier_c_price, base_min_price] if x is not None]
                    effective_min_price = min(candidates) if candidates else None
                valid_promotions: list[PromotionRule] = []
                promos_applied = item.get("promotion_applied", [])
                for promo_name in promos_applied:
                    matched_promotions = active_promo_names.get(promo_name) or []
                    if not matched_promotions:
                        review["status"] = "Exception"
                        review["risk_flags"].append(f"促销不存在或未生效：{promo_name}")
                    elif not any(promotion.sku_uuid == sku.id for promotion in matched_promotions):
                        review["status"] = "Exception"
                        review["risk_flags"].append(f"促销不适用于 SKU {sku_id}：{promo_name}")
                    else:
                        sku_promotions = [promotion for promotion in matched_promotions if promotion.sku_uuid == sku.id]
                        if len(sku_promotions) > 1:
                            review["status"] = "Exception"
                            review["risk_flags"].append(f"SKU {sku_id} 的促销规则重复：{promo_name}")
                        else:
                            valid_promotions.extend(sku_promotions)
                for promotion in valid_promotions:
                    discounted_min = promotion_discounted_min_price(effective_min_price, promotion)
                    if discounted_min is not None:
                        effective_min_price = discounted_min
                if effective_min_price is not None and unit_price < effective_min_price:
                    review["status"] = "Exception"
                    price_label = "促销最低价" if valid_promotions or is_promo_period else "最低限价"
                    review["risk_flags"].append(f"销售单价 {format_cents(unit_price)} 低于{price_label} {format_cents(effective_min_price)}")
                if pricing.max_price is not None and unit_price > pricing.max_price:
                    review["status"] = "Exception"
                    review["risk_flags"].append(f"销售单价 {format_cents(unit_price)} 高于最高限价 {format_cents(pricing.max_price)}")
            else:
                review["status"] = "Warning"
                review["risk_flags"].append(f"渠道 {channel} 未配置价格规则，无法完成价格审查")
        elif require_unit_price:
            review["status"] = "Warning"
            review["risk_flags"].append(f"SKU {sku_id} 未识别到销售单价，无法完成价格审查")
                
        if review["risk_flags"]:
            review["status"] = "Exception"

        item["review"] = review
        results.append(item)
        
    return results


def format_cents(value: int | None) -> str:
    if value is None:
        return "未配置"
    return f"{value / 100:.2f}元"

def extract_products_with_llm(session: Session, product_summary: str) -> list[dict]:
    """
    Use LLM to extract a list of materials from the material summary text.
    Expected output format is a JSON array of objects with keys:
    - sku_id (string)
    - unit_price (integer, representing cents)
    - promotion_applied (list of strings)
    """
    from backend.app.services.model_provider import call_model, extract_chat_content
    from backend.app.services.llm_fallback import active_model_config
    
    if not product_summary or not product_summary.strip():
        return []

    system_prompt = (
        "You are a helpful assistant that extracts material information from order text. "
        "Extract the materials into a JSON array. Each object MUST have: "
        "'sku_id' (string, SKU code if found; otherwise empty string), "
        "'sku_code' (string, same as sku_id if found), "
        "'unit_price' (integer, the price in cents, if not found use null), "
        "'promotion_applied' (list of strings, any discount or promo codes mentioned). "
        "Return ONLY valid JSON. No markdown formatting or extra text."
    )
    user_prompt = f"Extract materials from the following text:\n\n{product_summary}"
    
    try:
        config = active_model_config(session)
        if config is None:
            return []
        response = call_model(
            session,
            config,
            task_type="ProductPriceReviewExtract",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )
        content = extract_chat_content(response)
        # Clean up markdown if model still returns it
        if content.startswith("```json"):
            content = content[7:]
        if content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
            
        items = loads(content, [])
        if isinstance(items, list):
            return [normalize_extracted_product_item(item) for item in items if isinstance(item, dict)]
    except Exception as e:
        logger.warning("Failed to extract products with LLM: %s", e)
    
    return []


def semantic_match_skus(session: Session, query: str, limit: int = 5) -> list[ProductSKU]:
    """
    使用大语言模型（LLM）对 CRM 产品名称进行语义匹配，在 active 的 SKU 中搜寻对应的货品/物料。
    """
    from backend.app.services.model_provider import call_model, extract_chat_content
    from backend.app.services.llm_fallback import active_model_config
    
    query = str(query or "").strip()
    if not query:
        return []
        
    active_skus = (
        session.query(ProductSKU)
        .join(ProductSPU, ProductSKU.spu_uuid == ProductSPU.id)
        .filter(ProductSKU.status == "Active")
        .all()
    )
    if not active_skus:
        return []
        
    candidates = []
    for sku in active_skus:
        spu = sku.spu
        aliases = spu_review_aliases(spu) if spu else []
        candidates.append({
            "sku_id": sku.sku_id,
            "spu_name": spu.name if spu else "",
            "model": sku.model or "",
            "brand": spu.brand if spu else "",
            "category": spu.category if spu else "",
            "aliases": aliases
        })
        
    if len(candidates) > 30:
        scored = []
        for cand in candidates:
            score = 0
            q_norm = query.lower()
            if q_norm in cand["sku_id"].lower():
                score += 50
            if cand["spu_name"] and q_norm in cand["spu_name"].lower():
                score += 40
            if cand["model"] and q_norm in cand["model"].lower():
                score += 30
            for alias in cand["aliases"]:
                if q_norm in alias.lower() or alias.lower() in q_norm:
                    score += 40
            scored.append((cand, score))
        scored.sort(key=lambda x: x[1], reverse=True)
        candidates = [x[0] for x in scored[:30]]
        
    system_prompt = (
        "你是一个主数据专家。用户输入了一个来自 CRM 系统中的产品/物料/货品名称，你需要从以下给定的标准物料（候选）列表中，"
        "寻找在语义上最匹配的最多 5 个标准物料的 sku_id。\n"
        "注意：CRM 中的命名可能包含拼写差异、别名、缩写或中英文对照。请进行深度的语义匹配和型号比对。\n"
        "请务必只返回一个标准的 JSON 数组，数组中包含匹配到的 sku_id 字符串（按匹配度从高到低排序）。"
        "例如：[\"SKU-3D-SCANNER-PRO\", \"SKU-3D-SCANNER-LITE\"]\n"
        "如果没有任何合理的匹配，请返回 []。不要包含任何 markdown 块、不要包含 ```json 等代码包裹符，仅输出纯 JSON 数组文本。"
    )
    
    candidates_json = json.dumps(candidates, ensure_ascii=False)
    user_prompt = f"用户输入的 CRM 产品名称: {query}\n\n候选标准物料列表:\n{candidates_json}"
    
    try:
        config = active_model_config(session)
        if config is None:
            logger.warning("LLM semantic match skipped: model not configured")
            return []
            
        response = call_model(
            session,
            config,
            task_type="ProductSemanticMatch",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )
        content = extract_chat_content(response).strip()
        
        if content.startswith("```json"):
            content = content[7:]
        if content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        content = content.strip()
        
        matched_ids = loads(content, [])
        if not isinstance(matched_ids, list):
            return []
            
        sku_map = {sku.sku_id: sku for sku in active_skus}
        result_skus = []
        for sku_id in matched_ids:
            if sku_id in sku_map:
                result_skus.append(sku_map[sku_id])
        return result_skus
    except Exception as exc:
        logger.exception("LLM semantic product matching failed: %s", exc)
        return []


def preview_alias_import_from_excel(file_path: str, session: Session) -> dict:
    """预览 CRM 产品导入模板中的别名导入结果。

    模板格式（来自纷享销客）：
      产品名称 | 规格型号 | ... | 产品编码 | ...
    产品编码 → ProductSPU.spu_id
    产品名称 + 规格型号 → 预审别名
    """
    import openpyxl
    from backend.app.services.jsonutil import dumps

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb["产品导入模版"] if "产品导入模版" in wb.sheetnames else wb.active

    rows_iter = sheet.iter_rows(values_only=True)
    # Skip header row and example row
    next(rows_iter, None)  # header
    next(rows_iter, None)  # example/description

    items: list[dict] = []
    spu_cache: dict[str, ProductSPU | None] = {}
    matched = 0
    skipped_no_code = 0
    skipped_no_spu = 0
    skipped_no_alias = 0

    for row in rows_iter:
        if not row or not any(row):
            continue
        product_name = str(row[0] or "").strip() if len(row) > 0 else ""
        specification = str(row[1] or "").strip() if len(row) > 1 else ""
        product_code = str(row[3] or "").strip() if len(row) > 3 else ""  # 产品编码

        if not product_code:
            skipped_no_code += 1
            continue

        # Generate aliases: product_name and specification are both useful as search terms
        new_aliases_raw = []
        if product_name:
            new_aliases_raw.append(product_name)
        if specification and specification != product_name:
            new_aliases_raw.append(specification)

        if not new_aliases_raw:
            skipped_no_alias += 1
            continue

        # Look up SPU by spu_id (product code)
        if product_code not in spu_cache:
            spu = session.query(ProductSPU).filter(ProductSPU.spu_id == product_code).first()
            spu_cache[product_code] = spu
        spu = spu_cache[product_code]

        if spu is None:
            skipped_no_spu += 1
            continue

        matched += 1
        existing_aliases = spu_review_aliases(spu)
        all_merged = normalize_product_review_aliases(existing_aliases + new_aliases_raw)
        new_count = len(all_merged) - len(existing_aliases)

        items.append({
            "spu_uuid": spu.id,
            "spu_id": spu.spu_id,
            "product_name": spu.name,
            "existing_aliases": list(existing_aliases),
            "new_aliases": [a for a in new_aliases_raw if a not in existing_aliases],
            "merged_aliases": list(all_merged),
            "new_count": new_count,
        })

    wb.close()

    return {
        "total_rows": len(items),
        "matched": matched,
        "skipped_no_code": skipped_no_code,
        "skipped_no_spu": skipped_no_spu,
        "skipped_no_alias": skipped_no_alias,
        "items": items,
        "summary": {
            "spus_with_new": sum(1 for i in items if i["new_count"] > 0),
            "total_new_aliases": sum(i["new_count"] for i in items),
        },
    }


def confirm_alias_import_from_excel(preview_data: dict, session: Session) -> dict:
    """确认导入，将预览结果写入 SPU 别名（绕过 finished inventory 限制）。"""
    items = preview_data.get("items", [])
    updated = 0
    errors = 0
    error_details: list[str] = []

    for item in items:
        spu_uuid = item.get("spu_uuid")
        if not spu_uuid:
            continue
        try:
            spu = session.get(ProductSPU, spu_uuid)
            if spu is None:
                errors += 1
                error_details.append(f"{item.get('spu_id', '?')}: SPU 不存在")
                continue
            # Merge new aliases with existing
            existing = item.get("existing_aliases", [])
            new_raw = item.get("new_aliases", [])
            merged = normalize_product_review_aliases(existing + new_raw)
            info = loads(spu.extended_info_json, {}) if spu.extended_info_json else {}
            if not isinstance(info, dict):
                info = {}
            info["review_aliases"] = merged
            spu.extended_info_json = json.dumps(info, ensure_ascii=False)
            spu.updated_at = now_utc()
            updated += 1
        except Exception as exc:
            errors += 1
            error_details.append(f"{item.get('spu_id', '?')}: {exc}")

    if errors:
        # 部分失败不影响已成功的写入
        session.flush()

    return {
        "updated": updated,
        "errors": errors,
        "error_details": error_details[:10],
    }
