"""海外库存Excel导入服务

处理形如 海外库存总表-20260612.xlsx 的库存Excel，支持：
- 解析多仓库（美西仓库、Amazon US、Amazon DE、Amazon JP、Amazon UK、德国仓库等）
- 记录导入历史（文件名、时间、仓库、行数）
- 归档源文件到 data/inventory_archives/
- 写入快照到 ProductInventorySnapshot + InventorySnapshotHistory
- 查询库存变化走势
"""

from __future__ import annotations

import json
import os
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.app.models import (
    InventoryImportRecord,
    InventorySnapshotHistory,
    ProductInventorySnapshot,
    now_utc,
)


ARCHIVE_DIR = "data/inventory_archives"


def parse_inventory_excel(file_path: str) -> dict[str, Any]:
    """解析库存Excel，返回按仓库分组的数据

    Excel 列说明（基于 海外库存总表-20260612.xlsx）：
      B: SKU编码 / 物料代码
      C: 仓库名称（如 美西仓库、Amazon US、Amazon DE 等）
      D: 料号代码
      E: 中文品名（可能为 VLOOKUP 公式，用 data_only=True 获取缓存值）
      I: 库存数量（可能为 VLOOKUP 公式，用 data_only=True 获取缓存值）
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    if ws is None:
        return {"ok": False, "error": "无法读取工作表"}

    rows: list[dict[str, Any]] = []
    warehouse_counts: dict[str, int] = {}

    for ri in range(2, ws.max_row + 1):  # 从第2行开始（跳过标题行）
        sku = ws.cell(row=ri, column=2).value  # B列: SKU
        warehouse = ws.cell(row=ri, column=3).value  # C列: 仓库
        material_code = ws.cell(row=ri, column=4).value or sku  # D列: 料号代码
        name = ws.cell(row=ri, column=5).value  # E列: 中文品名
        qty = ws.cell(row=ri, column=9).value  # I列: 库存数量

        # 跳过空行
        if not sku and not material_code:
            continue

        # 处理 SKU（可能是数字，转为字符串）
        sku_str = str(sku).strip() if sku is not None else ""
        if not sku_str:
            sku_str = str(material_code).strip() if material_code else ""

        # 处理仓库名
        wh_str = str(warehouse).strip() if warehouse else ""

        # 处理品名（公式跳过，用料号代替）
        name_str = str(name).strip() if name and isinstance(name, str) and not name.startswith("=") else f"物料{sku_str}"
        if len(name_str) > 60:
            name_str = name_str[:60]

        # 处理数量
        try:
            qty_val = float(qty) if qty is not None else 0
        except (TypeError, ValueError):
            qty_val = 0

        row_data = {
            "sku": sku_str,
            "material_code": sku_str,
            "material_name": name_str,
            "warehouse": wh_str,
            "quantity": qty_val,
        }
        if wh_str:
            warehouse_counts[wh_str] = warehouse_counts.get(wh_str, 0) + 1
        rows.append(row_data)

    return {
        "ok": True,
        "file_name": Path(file_path).name,
        "total_rows": len(rows),
        "warehouse_counts": warehouse_counts,
        "warehouses": list(warehouse_counts.keys()),
        "rows": rows,
    }


def import_inventory_excel(
    session: Session,
    file_path: str,
    *,
    operated_by: str = "",
) -> dict[str, Any]:
    """导入库存Excel到系统

    流程：
      1. 解析Excel
      2. 创建导入记录
      3. 归档源文件
      4. 更新当前快照（ProductInventorySnapshot）
      5. 写入历史快照（InventorySnapshotHistory）
    """
    # 1. 解析
    result = parse_inventory_excel(file_path)
    if not result.get("ok"):
        return result

    rows = result["rows"]
    if not rows:
        return {"ok": False, "error": "Excel 中无有效数据行"}

    wh_counts = result["warehouse_counts"]
    ts = now_utc()

    # 2. 创建导入记录
    record = InventoryImportRecord(
        file_name=result["file_name"],
        warehouse=json.dumps(list(wh_counts.keys()), ensure_ascii=False),
        row_count=len(rows),
        status="Completed",
        operated_by=operated_by or "system",
        notes=f"仓库: {', '.join(wh_counts.keys())}, 总计{len(rows)}行",
    )
    session.add(record)
    session.flush()

    # 3. 归档源文件
    try:
        archive_dir = Path(ARCHIVE_DIR)
        archive_dir.mkdir(parents=True, exist_ok=True)
        dest = archive_dir / f"{date.today().isoformat()}_{Path(file_path).name}"
        shutil.copy2(file_path, dest)
    except Exception:
        pass  # 归档失败不影响主流程

    # 4. 更新当前快照 + 写入历史
    created_count = 0
    updated_count = 0
    for row in rows:
        wh = row["warehouse"]
        if not wh:
            continue
        mc = row["material_code"]
        mn = row["material_name"]
        qty = row["quantity"]

        # 更新或创建当前快照
        snap = session.query(ProductInventorySnapshot).filter(
            ProductInventorySnapshot.material_code == mc,
            ProductInventorySnapshot.warehouse_code == wh,
        ).first()

        if snap is None:
            snap = ProductInventorySnapshot(
                material_code=mc,
                warehouse_code=wh,
            )
            session.add(snap)
            created_count += 1
        else:
            updated_count += 1

        snap.material_name = mn
        snap.warehouse_name = wh
        snap.qty = qty
        snap.base_qty = qty
        snap.synced_at = ts
        snap.status = "Active"
        snap.updated_at = ts

        # 写入历史快照
        history = InventorySnapshotHistory(
            material_code=mc,
            material_name=mn,
            warehouse_code=wh,
            qty=qty,
            import_record_id=record.id,
            snapshot_date=ts,
        )
        session.add(history)

    session.flush()
    return {
        "ok": True,
        "record_id": record.id,
        "file_name": result["file_name"],
        "total_rows": len(rows),
        "created": created_count,
        "updated": updated_count,
        "warehouses": list(wh_counts.keys()),
    }


# ── 查询 API ──

def get_inventory_trends(
    session: Session,
    material_code: str = "",
    warehouse: str = "",
    days: int = 90,
) -> list[dict[str, Any]]:
    """查询指定物料/仓库的库存变化走势"""
    from datetime import timedelta

    since = now_utc() - timedelta(days=days)
    q = session.query(InventorySnapshotHistory).filter(InventorySnapshotHistory.created_at >= since)
    if material_code:
        q = q.filter(InventorySnapshotHistory.material_code == material_code)
    if warehouse:
        q = q.filter(InventorySnapshotHistory.warehouse_code == warehouse)
    q = q.order_by(InventorySnapshotHistory.warehouse_code, InventorySnapshotHistory.material_code, InventorySnapshotHistory.snapshot_date)
    return [
        {
            "material_code": h.material_code,
            "material_name": h.material_name,
            "warehouse_code": h.warehouse_code,
            "qty": h.qty,
            "snapshot_date": h.snapshot_date.isoformat() if h.snapshot_date else "",
            "import_record_id": h.import_record_id,
        }
        for h in q.all()
    ]


def list_import_records(
    session: Session,
    limit: int = 50,
) -> list[dict[str, Any]]:
    """列出最近的库存导入记录"""
    records = session.query(InventoryImportRecord).order_by(
        InventoryImportRecord.created_at.desc()
    ).limit(limit).all()
    return [
        {
            "id": r.id,
            "file_name": r.file_name,
            "warehouse": r.warehouse,
            "row_count": r.row_count,
            "status": r.status,
            "operated_by": r.operated_by,
            "notes": r.notes,
            "created_at": r.created_at.isoformat() if r.created_at else "",
        }
        for r in records
    ]
